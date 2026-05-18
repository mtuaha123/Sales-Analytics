import re
from io import BytesIO
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st

try:
    import pymupdf
except ImportError:
    import fitz as pymupdf

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt


# Optional OCR support for scanned / image-based PDFs
try:
    import numpy as np
    import cv2
    import pytesseract
    from PIL import Image as PILImage, ImageOps, ImageEnhance
    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False


from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
    Image,
)

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# -----------------------------
# Page config + custom styling
# -----------------------------
st.set_page_config(
    page_title="Sales Analytics",
    page_icon="📊",
    layout="wide"
)


st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', 'Segoe UI', Arial, sans-serif;
    }

    .stApp {
        background:
            radial-gradient(circle at 12% 8%, rgba(56,189,248,0.26) 0, rgba(56,189,248,0.02) 28%, transparent 34%),
            radial-gradient(circle at 86% 12%, rgba(168,85,247,0.22) 0, rgba(168,85,247,0.02) 25%, transparent 35%),
            linear-gradient(135deg, #020617 0%, #0f172a 48%, #111827 100%);
        color: #f8fafc;
    }

    .block-container {
        padding-top: 1.3rem;
        padding-bottom: 2.3rem;
        max-width: 1500px;
    }

    h1, h2, h3, h4, h5, h6 {
        color: #f8fafc !important;
        letter-spacing: -0.03em;
    }

    .hero-card {
        position: relative;
        overflow: hidden;
        background:
            linear-gradient(135deg, rgba(14,165,233,0.22), rgba(30,41,59,0.82)),
            linear-gradient(45deg, rgba(255,255,255,0.06), rgba(255,255,255,0.01));
        border: 1px solid rgba(226,232,240,0.18);
        border-radius: 28px;
        padding: 30px 32px;
        box-shadow: 0 24px 70px rgba(0,0,0,0.36), inset 0 1px 0 rgba(255,255,255,0.15);
        margin-bottom: 20px;
    }

    .hero-card::before {
        content: "";
        position: absolute;
        top: -80px;
        right: -80px;
        width: 260px;
        height: 260px;
        background: radial-gradient(circle, rgba(59,130,246,0.45), transparent 65%);
        filter: blur(3px);
    }

    .hero-title {
        position: relative;
        font-size: 2.35rem;
        font-weight: 850;
        color: #ffffff;
        margin-bottom: 8px;
        letter-spacing: -0.045em;
    }

    .hero-subtitle {
        position: relative;
        color: #dbeafe;
        font-size: 1.02rem;
        line-height: 1.65;
        max-width: 1050px;
    }

    .kpi-card {
        position: relative;
        overflow: hidden;
        background: linear-gradient(145deg, rgba(255,255,255,0.105), rgba(255,255,255,0.045));
        border: 1px solid rgba(226,232,240,0.16);
        border-radius: 22px;
        padding: 18px 20px;
        box-shadow: 0 16px 38px rgba(0,0,0,0.26), inset 0 1px 0 rgba(255,255,255,0.10);
        backdrop-filter: blur(14px);
        min-height: 130px;
        margin-bottom: 14px;
        transition: transform 0.18s ease, border-color 0.18s ease, box-shadow 0.18s ease;
    }

    .kpi-card:hover {
        transform: translateY(-3px);
        border-color: rgba(56,189,248,0.44);
        box-shadow: 0 20px 46px rgba(0,0,0,0.32), 0 0 0 1px rgba(56,189,248,0.10);
    }

    .kpi-card::after {
        content: "";
        position: absolute;
        width: 90px;
        height: 90px;
        right: -24px;
        bottom: -28px;
        border-radius: 999px;
        background: radial-gradient(circle, rgba(56,189,248,0.24), transparent 70%);
    }

    .kpi-label {
        font-size: 0.82rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: #93c5fd;
        margin-bottom: 10px;
        font-weight: 700;
    }

    .kpi-value {
        font-size: 1.68rem;
        line-height: 1.12;
        font-weight: 850;
        color: #ffffff;
        word-break: break-word;
    }

    .small-note {
        color: #cbd5e1;
        font-size: 0.82rem;
        margin-top: 10px;
        line-height: 1.45;
    }

    .summary-box {
        background: linear-gradient(135deg, rgba(15,23,42,0.82), rgba(30,41,59,0.66));
        border: 1px solid rgba(125,211,252,0.22);
        border-left: 6px solid #38bdf8;
        border-radius: 22px;
        padding: 20px 24px;
        color: #e2e8f0;
        box-shadow: 0 18px 42px rgba(0,0,0,0.28), inset 0 1px 0 rgba(255,255,255,0.08);
        line-height: 1.72;
        margin-top: 10px;
        margin-bottom: 22px;
        font-size: 1rem;
    }

    div[data-testid="stFileUploader"] {
        background: rgba(15,23,42,0.70);
        border: 1px dashed rgba(125,211,252,0.50);
        border-radius: 22px;
        padding: 16px;
        box-shadow: 0 14px 34px rgba(0,0,0,0.22);
    }

    div[data-testid="stFileUploader"] section {
        background: rgba(255,255,255,0.045);
        border-radius: 18px;
        border: 1px solid rgba(255,255,255,0.08);
    }

    .stButton > button,
    div[data-testid="stDownloadButton"] > button {
        background: linear-gradient(135deg, #2563eb, #06b6d4) !important;
        color: white !important;
        border: 0 !important;
        border-radius: 14px !important;
        padding: 0.70rem 1.05rem !important;
        font-weight: 800 !important;
        box-shadow: 0 14px 30px rgba(37,99,235,0.32) !important;
        transition: all 0.18s ease !important;
    }

    .stButton > button:hover,
    div[data-testid="stDownloadButton"] > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 18px 38px rgba(6,182,212,0.34) !important;
        filter: brightness(1.06);
    }

    div[data-testid="stAlert"] {
        border-radius: 18px;
        border: 1px solid rgba(255,255,255,0.12);
        box-shadow: 0 12px 28px rgba(0,0,0,0.20);
    }

    details {
        background: rgba(15,23,42,0.58) !important;
        border: 1px solid rgba(148,163,184,0.20) !important;
        border-radius: 18px !important;
        padding: 8px 12px !important;
        box-shadow: 0 10px 26px rgba(0,0,0,0.18);
    }

    summary {
        color: #e0f2fe !important;
        font-weight: 700 !important;
    }

    hr {
        border-color: rgba(148,163,184,0.22) !important;
        margin-top: 1.2rem !important;
        margin-bottom: 1.2rem !important;
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, rgba(15,23,42,0.98), rgba(2,6,23,0.98)) !important;
        border-right: 1px solid rgba(148,163,184,0.20);
    }

    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3 {
        color: #e0f2fe !important;
    }

    section[data-testid="stSidebar"] label {
        color: #cbd5e1 !important;
        font-weight: 650;
    }

    section[data-testid="stSidebar"] div[data-baseweb="select"] > div {
        background-color: rgba(255,255,255,0.065) !important;
        border: 1px solid rgba(125,211,252,0.22) !important;
        border-radius: 14px !important;
        color: #f8fafc !important;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 10px;
        background: rgba(255,255,255,0.035);
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 18px;
        padding: 8px;
    }

    .stTabs [data-baseweb="tab"] {
        background: rgba(255,255,255,0.055);
        border-radius: 14px;
        color: #cbd5e1;
        padding: 10px 18px;
        font-weight: 750;
        border: 1px solid rgba(255,255,255,0.06);
    }

    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #2563eb, #0891b2) !important;
        color: #ffffff !important;
        box-shadow: 0 12px 26px rgba(37,99,235,0.26);
    }

    div[data-testid="stDataFrame"] {
        background: rgba(15,23,42,0.72);
        border: 1px solid rgba(148,163,184,0.18);
        border-radius: 20px;
        padding: 10px;
        box-shadow: 0 18px 42px rgba(0,0,0,0.25);
        overflow: hidden;
    }

    div[data-testid="stPlotlyChart"] {
        background: linear-gradient(145deg, rgba(255,255,255,0.070), rgba(255,255,255,0.025));
        border: 1px solid rgba(226,232,240,0.12);
        border-radius: 22px;
        padding: 12px 12px 4px 12px;
        box-shadow: 0 18px 44px rgba(0,0,0,0.27), inset 0 1px 0 rgba(255,255,255,0.08);
        margin-bottom: 18px;
    }

    .stTextInput input,
    .stTextArea textarea {
        background: rgba(15,23,42,0.72) !important;
        border: 1px solid rgba(125,211,252,0.24) !important;
        border-radius: 14px !important;
        color: #f8fafc !important;
    }

    div[data-testid="stCaptionContainer"] {
        color: #cbd5e1 !important;
    }

    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)


st.markdown("""
<div class="hero-card">
    <div class="hero-title">📊 Sales Analytics</div>
    <div class="hero-subtitle">
        Upload invoice PDFs or invoice images and get a premium styled analytics dashboard with KPIs, filters, automatic summary, professional charts, styled Excel export, and PDF report with charts.
    </div>
</div>
""", unsafe_allow_html=True)


# -----------------------------
# Utility functions
# -----------------------------
def normalize_spaces(value):
    if value is None:
        return ""
    value = str(value).replace("\xa0", " ").replace("\u200b", "")
    value = re.sub(r"[ \t]+", " ", value)
    return value.strip()


def to_number(value):
    if value is None:
        return 0.0
    value = str(value).replace(",", "").replace("Rs.", "").replace("Rs", "").replace("-", "").strip()
    if not value:
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def format_currency(x):
    try:
        return f"Rs. {float(x):,.0f}"
    except Exception:
        return "Rs. 0"


def format_number(x):
    try:
        return f"{float(x):,.0f}"
    except Exception:
        return "0"


def group_words_into_lines(page, y_tolerance=3.0):
    words = page.get_text("words")
    if not words:
        return []

    words = sorted(words, key=lambda w: (w[1], w[0]))
    grouped_lines = []
    current_line = []
    current_y = None

    for word in words:
        x0, y0, x1, y1, text, *_ = word
        if current_y is None:
            current_line = [word]
            current_y = y0
        elif abs(y0 - current_y) <= y_tolerance:
            current_line.append(word)
            current_y = (current_y * 0.8) + (y0 * 0.2)
        else:
            line_text = " ".join(w[4] for w in sorted(current_line, key=lambda x: x[0]))
            grouped_lines.append(normalize_spaces(line_text))
            current_line = [word]
            current_y = y0

    if current_line:
        line_text = " ".join(w[4] for w in sorted(current_line, key=lambda x: x[0]))
        grouped_lines.append(normalize_spaces(line_text))

    return [line for line in grouped_lines if line]


def extract_pdf_content(uploaded_file):
    pdf_bytes = uploaded_file.getvalue()
    doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
    full_text_parts = []
    visual_lines = []

    for page in doc:
        full_text_parts.append(page.get_text("text", sort=True))
        visual_lines.extend(group_words_into_lines(page))

    return "\n".join(full_text_parts), visual_lines


def cut_before_delivery_list(text):
    parts = re.split(r"\bDelivery\s+LIST\b", text, flags=re.IGNORECASE)
    return parts[0] if parts else text


def extract_metadata(text, file_name):
    invoice_text = cut_before_delivery_list(text)

    date_match = re.search(r"\bDate\s+(\d{4}[/-]\d{1,2}[/-]\d{1,2})", invoice_text, re.IGNORECASE)
    due_match = re.search(r"\bDue\s+date\s+(\d{4}[/-]\d{1,2}[/-]\d{1,2})", invoice_text, re.IGNORECASE)
    invoice_match = re.search(r"\bINVOICE\s*#\s*\[?([^\]\n\r]+)\]?", invoice_text, re.IGNORECASE)
    subtotal_match = re.search(r"\bSubtotal\s+([\d,]+(?:\.\d+)?)", invoice_text, re.IGNORECASE)
    freight_match = re.search(r"\bFreight\s+Fee\s+([\d,]+(?:\.\d+)?)", invoice_text, re.IGNORECASE)
    balance_match = re.search(r"\bBalance\s+Rs\.?\s*([\d,]+(?:\.\d+)?)", invoice_text, re.IGNORECASE)

    benefit_values = re.findall(r"\bben(?:e|i)fit\s+([\d,]+(?:\.\d+)?)", invoice_text, re.IGNORECASE)
    benefit_total = sum(to_number(v) for v in benefit_values)

    customer_details = ""
    customer_match = re.search(
        r"Seller\s+CUSTOMER\s+(.*?)\s+Phone\s*:",
        invoice_text,
        re.IGNORECASE | re.DOTALL
    )
    if customer_match:
        customer_details = normalize_spaces(customer_match.group(1))

    return {
        "File Name": file_name,
        "Invoice No": normalize_spaces(invoice_match.group(1)) if invoice_match else file_name,
        "Invoice Date": date_match.group(1) if date_match else "",
        "Due Date": due_match.group(1) if due_match else "",
        "Customer Details": customer_details,
        "Invoice Subtotal": to_number(subtotal_match.group(1)) if subtotal_match else 0.0,
        "Benefit / Discount": benefit_total,
        "Freight Fee": to_number(freight_match.group(1)) if freight_match else 0.0,
        "Invoice Balance": to_number(balance_match.group(1)) if balance_match else 0.0,
    }


def clean_product_description(description):
    description = normalize_spaces(description)

    match = re.match(
        r"^(?P<product>.+?)\s+Rs\.?\s*(?P<retail_price>\d+(?:\.\d+)?)\s+(?P<packing>\d+\s*\*\s*\d+\s*\*\s*\d+)$",
        description,
        re.IGNORECASE
    )
    if match:
        product = normalize_spaces(match.group("product"))
        retail_price = to_number(match.group("retail_price"))
        packing = normalize_spaces(match.group("packing")).replace(" ", "")
        return product, retail_price, packing

    match_2 = re.match(
        r"^(?P<product>.+?)\s+Rs\.?\s*(?P<retail_price>\d+(?:\.\d+)?)",
        description,
        re.IGNORECASE
    )
    if match_2:
        return normalize_spaces(match_2.group("product")), to_number(match_2.group("retail_price")), ""

    return description, 0.0, ""


def make_row(match_dict, metadata, parser_method):
    raw_description = normalize_spaces(match_dict["description"])
    product, retail_price, packing = clean_product_description(raw_description)
    carton_price = to_number(match_dict["unit_price"])
    quantity = int(to_number(match_dict["quantity"]))
    amount = to_number(match_dict["amount"])
    calculated_amount = carton_price * quantity
    difference = round(amount - calculated_amount, 2)

    return {
        **metadata,
        "Sr No": int(match_dict["sr_no"]),
        "Product Raw": raw_description,
        "Product": product,
        "Retail Price": retail_price,
        "Packing": packing,
        "Carton / Unit Price": carton_price,
        "Quantity": quantity,
        "Amount": amount,
        "Calculated Amount": calculated_amount,
        "Amount Difference": difference,
        "Parser Method": parser_method,
    }


def parse_rows_from_joined_text(text, metadata):
    invoice_text = cut_before_delivery_list(text)

    header_match = re.search(
        r"No\.\s*DESCRIPTION\s+UNIT\s+PRICE\s+QTY\s+AMOUNT",
        invoice_text,
        re.IGNORECASE
    )
    if header_match:
        invoice_text = invoice_text[header_match.end():]

    end_match = re.search(
        r"(CUSTOMER\s+SIGNATURE|I hereby|Subtotal\s+[\d,]+)",
        invoice_text,
        re.IGNORECASE
    )
    if end_match:
        invoice_text = invoice_text[:end_match.start()]

    joined = normalize_spaces(invoice_text.replace("\n", " "))

    row_pattern = re.compile(
        r"(?<!\S)"
        r"(?P<sr_no>\d{1,3})\s+"
        r"(?P<description>.+?)\s+"
        r"(?P<unit_price>\d{1,3}(?:,\d{3})*\.\d{2})\s+"
        r"(?P<quantity>\d+(?:\.\d+)?)\s+"
        r"(?P<amount>\d{1,3}(?:,\d{3})*(?:\.\d+)?)"
        r"(?=\s+\d{1,3}\s+[A-Za-z(]|\s+CUSTOMER|\s+Subtotal|\s*$)",
        re.IGNORECASE
    )

    rows = []
    for match in row_pattern.finditer(joined):
        rows.append(make_row(match.groupdict(), metadata, "joined_text_regex"))

    return rows


def parse_rows_from_visual_lines(visual_lines, metadata):
    useful_lines = []
    inside_table = False

    for line in visual_lines:
        clean_line = normalize_spaces(line)

        if re.search(r"\bDelivery\s+LIST\b", clean_line, re.IGNORECASE):
            break

        if re.search(r"No\.\s*DESCRIPTION\s+UNIT\s+PRICE\s+QTY\s+AMOUNT", clean_line, re.IGNORECASE):
            inside_table = True
            continue

        if inside_table:
            if re.search(r"(CUSTOMER\s+SIGNATURE|I hereby|Subtotal\s+[\d,]+)", clean_line, re.IGNORECASE):
                break
            useful_lines.append(clean_line)

    line_pattern = re.compile(
        r"^"
        r"(?P<sr_no>\d{1,3})\s+"
        r"(?P<description>.+?)\s+"
        r"(?P<unit_price>\d{1,3}(?:,\d{3})*\.\d{2})\s+"
        r"(?P<quantity>\d+(?:\.\d+)?)\s+"
        r"(?P<amount>\d{1,3}(?:,\d{3})*(?:\.\d+)?)"
        r"$",
        re.IGNORECASE
    )

    rows = []
    for line in useful_lines:
        match = line_pattern.match(line)
        if match:
            rows.append(make_row(match.groupdict(), metadata, "visual_line_regex"))

    return rows


def deduplicate_rows(rows):
    seen = set()
    clean_rows = []
    for row in rows:
        key = (
            row.get("File Name"),
            row.get("Invoice No"),
            row.get("Sr No"),
            row.get("Product"),
            row.get("Quantity"),
            round(row.get("Amount", 0), 2),
        )
        if key not in seen:
            seen.add(key)
            clean_rows.append(row)
    return clean_rows


def parse_invoice_pdf(uploaded_file):
    text, visual_lines = extract_pdf_content(uploaded_file)
    metadata = extract_metadata(text, uploaded_file.name)

    rows = []
    rows.extend(parse_rows_from_joined_text(text, metadata))
    rows.extend(parse_rows_from_visual_lines(visual_lines, metadata))
    rows = deduplicate_rows(rows)

    # Important: scanned/image-based PDFs usually extract only text like "CamScanner".
    # If no invoice rows are found, run OCR fallback.
    ocr_message = ""
    if not rows:
        ocr_rows, ocr_message = parse_invoice_pdf_with_ocr(uploaded_file, metadata)
        if ocr_rows:
            rows = ocr_rows

    return rows, text, visual_lines, ocr_message


def is_image_file(uploaded_file):
    """
    Returns True for direct image uploads.
    """
    file_name = uploaded_file.name.lower()
    image_extensions = (".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff")
    return file_name.endswith(image_extensions)


def parse_invoice_image(uploaded_file):
    """
    Parses direct image uploads using the same OCR logic used for scanned PDFs.
    Supported image types: PNG, JPG, JPEG, WEBP, BMP, TIFF.
    """
    if not OCR_AVAILABLE:
        return (
            [],
            "",
            [],
            "OCR libraries are not installed. Install pytesseract, pillow, opencv-python-headless and tesseract-ocr."
        )

    try:
        uploaded_file.seek(0)
        image = PILImage.open(uploaded_file).convert("RGB")

        # OCR the full image once to extract available metadata such as invoice/date/subtotal.
        image_text = ocr_single_image_text(image, config="--oem 3 --psm 6")
        metadata = extract_metadata(image_text, uploaded_file.name)

        # For direct images, invoice number may not be captured by OCR.
        # Use file name as fallback so filters still work properly.
        if not metadata.get("Invoice No") or metadata.get("Invoice No") == uploaded_file.name:
            metadata["Invoice No"] = Path(uploaded_file.name).stem if "Path" in globals() else uploaded_file.name

        rows = []
        rows.extend(parse_rows_from_ocr_grid(image, metadata))

        # Full-page OCR fallback if grid detection does not find rows.
        if not rows:
            rows.extend(parse_rows_from_ocr_full_page(image, metadata))

        rows = deduplicate_rows_fuzzy(rows)

        return rows, image_text, [], ""

    except Exception as exc:
        return [], "", [], f"Image OCR failed: {exc}"


def parse_uploaded_invoice_file(uploaded_file):
    """
    Unified parser:
    - Text-based PDF: direct PDF text/table parsing
    - Scanned PDF: OCR fallback
    - Direct image upload: OCR parser
    """
    if is_image_file(uploaded_file):
        return parse_invoice_image(uploaded_file)

    return parse_invoice_pdf(uploaded_file)


def kpi_card(label, value, note=""):
    st.markdown(
        f"""
        <div class="kpi-card">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{value}</div>
            <div class="small-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True
    )


def styled_bar_chart(df, x, y, title):
    fig = px.bar(df, x=x, y=y, text=y, title=title, template="plotly_dark")
    fig.update_traces(texttemplate="%{text:,.0f}", textposition="outside")
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(255,255,255,0.02)",
        font=dict(color="white"),
        margin=dict(l=20, r=20, t=50, b=20),
        title_x=0.02,
        xaxis_title=None,
        yaxis_title=None
    )
    return fig


def styled_pie_chart(df, names, values, title):
    fig = px.pie(df, names=names, values=values, hole=0.45, title=title, template="plotly_dark")
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="white"),
        margin=dict(l=20, r=20, t=50, b=20),
        title_x=0.02
    )
    fig.update_traces(textinfo="percent+label")
    return fig


def styled_line_chart(df, x, y, title):
    fig = px.line(df, x=x, y=y, markers=True, title=title, template="plotly_dark")
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(255,255,255,0.02)",
        font=dict(color="white"),
        margin=dict(l=20, r=20, t=50, b=20),
        title_x=0.02,
        xaxis_title=None,
        yaxis_title=None
    )
    return fig


def build_product_summary(filtered_df):
    return (
        filtered_df.groupby("Product", as_index=False)
        .agg(
            Quantity=("Quantity", "sum"),
            Amount=("Amount", "sum"),
            Invoices=("Invoice No", "nunique"),
        )
        .sort_values("Amount", ascending=False)
    )


def build_invoice_summary(filtered_df):
    return (
        filtered_df.groupby(["Invoice No", "Invoice Date"], as_index=False)
        .agg(
            Quantity=("Quantity", "sum"),
            Amount=("Amount", "sum"),
            Invoice_Subtotal=("Invoice Subtotal", "max"),
            Invoice_Balance=("Invoice Balance", "max"),
            Freight_Fee=("Freight Fee", "max"),
            Benefit_Discount=("Benefit / Discount", "max"),
        )
        .sort_values("Invoice Date")
    )


def generate_auto_summary(filtered_df, product_summary, invoice_summary, monthly_sales):
    total_sales = filtered_df["Amount"].sum()
    total_qty = filtered_df["Quantity"].sum()
    total_invoices = filtered_df["Invoice No"].nunique()
    total_products = filtered_df["Product"].nunique()

    top_revenue = product_summary.sort_values("Amount", ascending=False).iloc[0]
    top_quantity = product_summary.sort_values("Quantity", ascending=False).iloc[0]
    bottom_quantity = product_summary[product_summary["Quantity"] > 0].sort_values("Quantity", ascending=True).iloc[0]
    bottom_revenue = product_summary[product_summary["Amount"] > 0].sort_values("Amount", ascending=True).iloc[0]

    top5_share = product_summary.head(5)["Amount"].sum() / total_sales * 100 if total_sales else 0
    top_revenue_share = top_revenue["Amount"] / total_sales * 100 if total_sales else 0

    total_discount = filtered_df.groupby("Invoice No")["Benefit / Discount"].max().sum()
    total_freight = filtered_df.groupby("Invoice No")["Freight Fee"].max().sum()

    month_sentence = ""
    if len(monthly_sales) >= 2:
        last = monthly_sales.iloc[-1]
        previous = monthly_sales.iloc[-2]
        if previous["Amount"] != 0:
            growth = ((last["Amount"] - previous["Amount"]) / previous["Amount"]) * 100
            direction = "increased" if growth >= 0 else "decreased"
            month_sentence = (
                f" Sales {direction} by {abs(growth):.1f}% from {previous['Month']} to {last['Month']}."
            )
    elif len(monthly_sales) == 1:
        month_sentence = f" The available data relates to {monthly_sales.iloc[0]['Month']}."

    summary = (
        f"{top_revenue['Product']} remained the highest revenue product with {format_currency(top_revenue['Amount'])}, "
        f"contributing approximately {top_revenue_share:.1f}% of total sales. "
        f"The dashboard covers {format_currency(total_sales)} sales, {format_number(total_qty)} quantity, "
        f"{total_invoices} invoice(s), and {total_products} unique product(s). "
        f"{top_quantity['Product']} was the highest quantity-selling item with {format_number(top_quantity['Quantity'])} units/cartons. "
        f"The top 5 products contributed around {top5_share:.1f}% of total sales, showing the main revenue concentration. "
        f"On the low-selling side, {bottom_quantity['Product']} recorded the lowest quantity movement "
        f"with {format_number(bottom_quantity['Quantity'])} units/cartons, while {bottom_revenue['Product']} generated the lowest revenue at "
        f"{format_currency(bottom_revenue['Amount'])}. "
        f"Total discount/benefit stood at {format_currency(total_discount)}, while total freight fee stood at {format_currency(total_freight)}."
        f"{month_sentence}"
    )
    return summary


def dataframe_for_display(df, money_cols=None, number_cols=None):
    display_df = df.copy()
    money_cols = money_cols or []
    number_cols = number_cols or []

    for col in money_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(format_currency)

    for col in number_cols:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(format_number)

    return display_df




# -----------------------------
# OCR fallback for scanned / image-based PDFs
# -----------------------------
def standardize_product_name(product_text):
    """
    OCR often distorts product names. This function normalizes common product names
    so the dashboard does not show duplicate/misspelled products.
    """
    text = normalize_spaces(product_text)
    low = text.lower()

    replacements = {
        "crispy": "Crispy cookies",
        "etispy": "Crispy cookies",
        "crsp": "Crispy cookies",
        "breakfast": "Breakfast Milk biscuits",
        "milk biscuit": "Breakfast Milk biscuits",
        "strawberry": "Vanilla Strawberry",
        "strawber": "Vanilla Strawberry",
        "stewie": "Vanilla Strawberry",
        "lemon": "Vanilla Lemon",
        "teron": "Vanilla Lemon",
        "temon": "Vanilla Lemon",
        "classic": "Classic supreme",
        "supreme": "Classic supreme",
        "zoo": "Children Zoo",
        "star": "Children Star",
        "cheese": "Children Cheese",
        "good taste": "Good Taste",
    }

    for key, value in replacements.items():
        if key in low:
            return value

    # Remove obvious OCR noise from start/end while keeping unknown product names usable
    text = re.sub(r"^[^A-Za-z]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def clean_ocr_line(line):
    line = str(line)
    line = line.replace("|", " ")
    line = line.replace("[", " ").replace("]", " ")
    line = line.replace("_", " ").replace("—", " ").replace("~", " ")
    line = line.replace('"', "*").replace("°", "*").replace("×", "*")
    line = re.sub(r"\b(?:R5|RS|rs|ns|Bs|B5)\.?\s*", "Rs.", line)
    line = re.sub(r"\s+", " ", line).strip()
    return line


def ocr_number_to_float(value):
    """
    Converts OCR number text to float. Handles common OCR mistakes.
    """
    if value is None:
        return 0.0

    value = str(value)
    value = value.translate(str.maketrans({
        "O": "0", "o": "0",
        "I": "1", "l": "1",
        "S": "5", "s": "5",
        "B": "8",
    }))

    value = re.sub(r"[^0-9,.\-]", "", value)

    if value.count(".") > 1:
        parts = value.split(".")
        value = "".join(parts[:-1]) + "." + parts[-1]

    value = value.strip()
    if not value:
        return 0.0

    try:
        return float(value.replace(",", ""))
    except Exception:
        return 0.0



KNOWN_UNIT_PRICE_RULES = [
    ("crispy", "1*24*18", 3240.00),
    ("crispy", "1*16*18", 4118.40),
    ("breakfast", "1*16*18", 4118.40),
    ("vanilla strawberry", "1*24*18", 3240.00),
    ("vanilla strawberry", "1*15*24", 2700.00),
    ("vanilla lemon", "1*24*18", 3240.00),
    ("classic supreme", "1*24*18", 3240.00),
    ("children zoo", "50pack", 1000.00),
    ("children star", "50pack", 1000.00),
    ("children cheese", "50pack", 1150.00),
    ("good taste", "1*12*18", 3088.80),
    ("sweet crispy cracker", "1*12*24", 4118.40),
    ("kids crush", "1*36*18", 4212.00),
    ("waha", "1*12*18", 3088.80),
]


def compact_for_matching(text):
    text = str(text).lower()
    text = text.replace("x", "*").replace("×", "*").replace(" ", "")
    text = text.replace("l", "1").replace("i", "1")
    text = text.replace("|", "1")
    return text


def known_unit_price_from_text(description):
    """
    Uses product + packing hints to recover unit price when OCR misses a numeric cell.
    This is only a fallback for scanned PDFs.
    """
    if not description:
        return 0.0

    low = str(description).lower()
    compact = compact_for_matching(description)

    for product_key, packing_key, unit_price in KNOWN_UNIT_PRICE_RULES:
        product_ok = product_key in low
        packing_ok = packing_key.replace(" ", "").lower() in compact
        if product_ok and packing_ok:
            return float(unit_price)

    # More forgiving fallbacks for common product names when packing is badly OCR'd
    if "children zoo" in low or ("children" in low and "zoo" in low):
        return 1000.00
    if "children star" in low or ("children" in low and "star" in low):
        return 1000.00
    if "children cheese" in low or ("children" in low and "cheese" in low):
        return 1150.00
    if "good taste" in low:
        return 3088.80
    if "classic" in low and "supreme" in low:
        return 3240.00

    return 0.0


def amount_qty_unit_correction(unit_price, quantity, amount):
    """
    Corrects common OCR numeric errors by enforcing invoice arithmetic:
    Amount = Unit Price × Quantity.
    Returns corrected unit, quantity, amount, and correction note.
    """
    note = []
    unit_price = float(unit_price or 0)
    amount = float(amount or 0)
    quantity = int(quantity or 0)

    if unit_price > 0 and amount > 0:
        inferred_qty = round(amount / unit_price)
        if 1 <= inferred_qty <= 10000:
            inferred_amount = unit_price * inferred_qty
            if abs(inferred_amount - amount) <= max(5, amount * 0.005):
                if quantity <= 0 or abs(quantity - inferred_qty) >= max(2, inferred_qty * 0.40):
                    note.append(f"Qty corrected from {quantity} to {inferred_qty}")
                    quantity = int(inferred_qty)

    if unit_price > 0 and quantity > 0:
        calculated = unit_price * quantity
        if amount <= 0:
            note.append(f"Amount filled from calculation: {calculated:.2f}")
            amount = calculated
        elif abs(calculated - amount) > max(5, amount * 0.05):
            # If the entered amount does not divide cleanly into a sensible quantity,
            # it is usually an OCR amount error like 276,800 instead of 226,800.
            inferred_qty_from_amount = round(amount / unit_price) if unit_price else 0
            amount_divides_cleanly = abs((unit_price * inferred_qty_from_amount) - amount) <= max(5, amount * 0.025)
            current_qty_reasonable = 1 <= quantity <= 10000

            if current_qty_reasonable and not amount_divides_cleanly:
                note.append(f"Amount corrected from {amount:.2f} to {calculated:.2f}")
                amount = calculated
            elif current_qty_reasonable and amount_divides_cleanly and abs(inferred_qty_from_amount - quantity) <= 2:
                quantity = int(inferred_qty_from_amount)
            elif current_qty_reasonable and amount > calculated * 0.5 and amount < calculated * 1.5:
                note.append(f"Amount corrected from {amount:.2f} to {calculated:.2f}")
                amount = calculated

    if unit_price <= 0 and quantity > 0 and amount > 0:
        inferred_unit = amount / quantity
        if 100 <= inferred_unit <= 100000:
            note.append(f"Unit price inferred as {inferred_unit:.2f}")
            unit_price = inferred_unit

    return unit_price, quantity, amount, "; ".join(note)


def make_ocr_row(metadata, sr_no, description, unit_price, quantity, amount, parser_method, raw_ocr=""):
    product, retail_price, packing = clean_product_description(description)
    product = standardize_product_name(product)

    raw_unit_price = float(unit_price or 0)
    raw_quantity = int(quantity or 0)
    raw_amount = float(amount or 0)

    # Recover unit price from product/packing when OCR misses or distorts the unit-price cell.
    known_unit = known_unit_price_from_text(description)
    if known_unit > 0 and (raw_unit_price <= 0 or abs(raw_unit_price - known_unit) > max(10, known_unit * 0.20)):
        unit_price = known_unit
    else:
        unit_price = raw_unit_price

    quantity = raw_quantity
    amount = raw_amount

    unit_price, quantity, amount, correction_note = amount_qty_unit_correction(unit_price, quantity, amount)

    calculated_amount = unit_price * quantity
    difference = round(amount - calculated_amount, 2)

    review_required = "Yes" if abs(difference) > max(5, float(amount or 0) * 0.05) else "No"
    if correction_note:
        review_required = "Review"

    return {
        **metadata,
        "Sr No": int(sr_no or 0),
        "Product Raw": normalize_spaces(description),
        "Product": product,
        "Retail Price": retail_price,
        "Packing": packing,
        "Carton / Unit Price": float(unit_price or 0),
        "Quantity": int(quantity or 0),
        "Amount": float(amount or 0),
        "Calculated Amount": calculated_amount,
        "Amount Difference": difference,
        "Parser Method": parser_method,
        "OCR Raw Text": raw_ocr,
        "OCR Raw Unit Price": raw_unit_price,
        "OCR Raw Quantity": raw_quantity,
        "OCR Raw Amount": raw_amount,
        "OCR Auto Correction": correction_note,
        "OCR Review Required": review_required,
    }

def parse_ocr_invoice_line(line, metadata, parser_method):
    """
    Parses one OCR line from scanned invoice table.
    Handles both full rows and partial rows where OCR captured only product + amount.
    """
    raw_line = line
    line = clean_ocr_line(line)
    low = line.lower()

    likely_product_words = [
        "cookie", "cookies", "biscuit", "biscuits", "vanilla", "children", "good", "taste",
        "supreme", "zoo", "star", "cheese", "crispy", "breakfast", "classic", "milk"
    ]
    if not any(word in low for word in likely_product_words):
        return None

    sr_no = 0
    sr_match = re.search(r"\b(\d{1,2})\b", line[:12])
    if sr_match:
        sr_no = int(sr_match.group(1))
        line_without_sr = line[sr_match.end():].strip()
    else:
        line_without_sr = line

    token_re = re.compile(r"(?<![A-Za-z])[-]?[0-9OolISsB,\.]{1,14}(?:\.[0-9OolISsB]{1,3})?(?![A-Za-z])")
    tokens = []
    for match in token_re.finditer(line_without_sr):
        raw_token = match.group(0)
        value = ocr_number_to_float(raw_token)
        if value > 0:
            tokens.append((match.start(), match.end(), raw_token, value))

    money_tokens = [
        t for t in tokens
        if ("," in t[2] or "." in t[2]) and t[3] >= 100
    ]

    description = ""
    unit_price = 0.0
    quantity = 0
    amount = 0.0

    if len(money_tokens) >= 2:
        unit_token = money_tokens[-2]
        amount_token = money_tokens[-1]
        unit_price = unit_token[3]
        amount = amount_token[3]

        qty_candidates = [
            t for t in tokens
            if t[0] >= unit_token[1] and t[1] <= amount_token[0] and 1 <= t[3] <= 10000
        ]
        between_unit_and_amount = line_without_sr[unit_token[1]:amount_token[0]].lower()
        if qty_candidates:
            quantity = int(round(qty_candidates[-1][3]))
        elif re.search(r"\b(g[aogq]|q[aogq]|go|g0|6o|o0)\b", between_unit_and_amount):
            # Tesseract often reads the quantity '60' as 'ga', 'go', 'g0', etc.
            quantity = 60
        elif unit_price > 0 and amount > 0:
            inferred_qty = round(amount / unit_price)
            if 1 <= inferred_qty <= 10000:
                quantity = int(inferred_qty)

        description = line_without_sr[:unit_token[0]].strip()

    elif len(money_tokens) == 1:
        # Common OCR partial row: "Children Star Rs.30 50Pack 20,000.00"
        amount_token = money_tokens[-1]
        amount = amount_token[3]
        description = line_without_sr[:amount_token[0]].strip()
        unit_price = known_unit_price_from_text(description)
        if unit_price > 0 and amount > 0:
            inferred_qty = round(amount / unit_price)
            if 1 <= inferred_qty <= 10000:
                quantity = int(inferred_qty)
        else:
            return None
    else:
        return None

    description = re.sub(r"^[^A-Za-z]+", "", description)
    description = normalize_spaces(description)

    # Remove obvious leading OCR noise before product names.
    description = re.sub(r"^(ari|fog|fag|r)\s+", "", description, flags=re.IGNORECASE).strip()

    if len(description) < 3 or amount <= 0:
        return None

    return make_ocr_row(
        metadata=metadata,
        sr_no=sr_no,
        description=description,
        unit_price=unit_price,
        quantity=quantity,
        amount=amount,
        parser_method=parser_method,
        raw_ocr=raw_line
    )

def render_pdf_pages_for_ocr(uploaded_file, zoom=4):
    """
    Converts PDF pages to PIL images for OCR.
    """
    pdf_bytes = uploaded_file.getvalue()
    doc = pymupdf.open(stream=pdf_bytes, filetype="pdf")
    images = []

    for page in doc:
        pix = page.get_pixmap(matrix=pymupdf.Matrix(zoom, zoom), alpha=False)
        img = PILImage.open(BytesIO(pix.tobytes("png"))).convert("RGB")
        images.append(img)

    return images


def detect_table_grid_positions(pil_img):
    """
    Detects strong table grid lines in scanned invoices.
    Returns x/y line positions for grid-based OCR.
    """
    img = np.array(pil_img)
    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)

    threshold = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_MEAN_C,
        cv2.THRESH_BINARY_INV,
        31, 15
    )

    horizontal_kernel = cv2.getStructuringElement(
        cv2.MORPH_RECT,
        (max(30, gray.shape[1] // 30), 1)
    )
    vertical_kernel = cv2.getStructuringElement(
        cv2.MORPH_RECT,
        (1, max(30, gray.shape[0] // 30))
    )

    horizontal_lines = cv2.morphologyEx(threshold, cv2.MORPH_OPEN, horizontal_kernel, iterations=1)
    vertical_lines = cv2.morphologyEx(threshold, cv2.MORPH_OPEN, vertical_kernel, iterations=1)

    horizontal_projection = np.sum(horizontal_lines > 0, axis=1)
    vertical_projection = np.sum(vertical_lines > 0, axis=0)

    y_positions = np.where(horizontal_projection > gray.shape[1] * 0.30)[0]
    x_positions = np.where(vertical_projection > gray.shape[0] * 0.20)[0]

    def merge_positions(positions, tolerance=20):
        if len(positions) == 0:
            return []
        groups = []
        current = [int(positions[0])]

        for position in positions[1:]:
            position = int(position)
            if position - current[-1] <= tolerance:
                current.append(position)
            else:
                groups.append(current)
                current = [position]

        groups.append(current)
        return [int(sum(group) / len(group)) for group in groups]

    x_lines = merge_positions(x_positions, tolerance=25)
    y_lines = merge_positions(y_positions, tolerance=25)

    return x_lines, y_lines


def ocr_single_image_text(pil_img, config="--psm 6"):
    gray = pil_img.convert("L")
    gray = ImageEnhance.Contrast(gray).enhance(1.5)
    return pytesseract.image_to_string(gray, config=config)


def parse_rows_from_ocr_full_page(pil_img, metadata):
    """
    Full-page OCR fallback. Useful when table grid detection is not perfect.
    """
    rows = []
    for config in ["--psm 6", "--psm 11"]:
        text = ocr_single_image_text(pil_img, config=config)
        for line in text.splitlines():
            row = parse_ocr_invoice_line(line, metadata, f"ocr_full_page_{config.replace(' ', '_')}")
            if row:
                rows.append(row)
    return rows


def parse_rows_from_ocr_grid(pil_img, metadata):
    """
    Faster grid-based OCR fallback.
    OCRs each table row once first, then retries/cell-combines only when needed.
    """
    rows = []
    x_lines, y_lines = detect_table_grid_positions(pil_img)

    if len(x_lines) < 5 or len(y_lines) < 4:
        return rows

    header_index = None
    for i in range(0, min(len(y_lines) - 1, 10)):
        crop = pil_img.crop((x_lines[0], y_lines[i], x_lines[-1], y_lines[i + 1]))
        txt = ocr_single_image_text(crop, config="--oem 3 --psm 7")
        if "description" in txt.lower() or "unit" in txt.lower() or "amount" in txt.lower():
            header_index = i
            break

    if header_index is None:
        header_index = 1

    # Process likely invoice rows only; stop after many blank rows.
    blank_streak = 0
    for i in range(header_index + 1, len(y_lines) - 1):
        y0, y1 = y_lines[i], y_lines[i + 1]
        if y1 - y0 < 18:
            continue

        row_crop = pil_img.crop((x_lines[0], y0, x_lines[-1], y1))
        row_crop = ImageOps.expand(row_crop, border=18, fill="white")

        candidate_rows = []

        # First pass: best speed/accuracy balance for table rows.
        row_text = ocr_single_image_text(row_crop, config="--oem 3 --psm 6")
        row = parse_ocr_invoice_line(row_text, metadata, "ocr_grid_row_psm6")
        if row:
            candidate_rows.append(row)

        # Retry only when first pass fails or is incomplete.
        needs_retry = not candidate_rows or any(
            candidate_rows[0].get(col, 0) in [0, 0.0, ""]
            for col in ["Carton / Unit Price", "Quantity", "Amount"]
        )
        if candidate_rows and abs(float(candidate_rows[0].get("Amount Difference", 0) or 0)) > 5:
            needs_retry = True
        if candidate_rows and candidate_rows[0].get("OCR Review Required") in ["Yes", "Review"]:
            needs_retry = True

        if needs_retry:
            row_text_7 = ocr_single_image_text(row_crop, config="--oem 3 --psm 7")
            row = parse_ocr_invoice_line(row_text_7, metadata, "ocr_grid_row_psm7")
            if row:
                candidate_rows.append(row)

        # Cell-combined OCR only for failed/incomplete rows.
        if needs_retry and len(x_lines) >= 6:
            cell_texts = []
            for j in range(0, 5):
                cell_crop = pil_img.crop((x_lines[j], y0, x_lines[j + 1], y1))
                cell_crop = ImageOps.expand(cell_crop, border=12, fill="white")
                cell_cfg = "--oem 3 --psm 7"
                if j in [0, 2, 3, 4]:
                    cell_cfg += " -c tessedit_char_whitelist=0123456789.,"
                cell_texts.append(ocr_single_image_text(cell_crop, config=cell_cfg).replace("\n", " "))
            combined = " ".join(cell_texts)
            row = parse_ocr_invoice_line(combined, metadata, "ocr_grid_cells_combined")
            if row:
                candidate_rows.append(row)

        if candidate_rows:
            blank_streak = 0

            def row_score(row):
                score = 0
                score += 5 if row.get("Product") else 0
                score += 4 if row.get("Carton / Unit Price", 0) > 0 else 0
                score += 4 if row.get("Quantity", 0) > 0 else 0
                score += 4 if row.get("Amount", 0) > 0 else 0
                score += 4 if abs(float(row.get("Amount Difference", 0) or 0)) <= 5 else 0
                score += 3 if row.get("OCR Review Required") == "No" else 0
                score -= min(abs(float(row.get("Amount Difference", 0) or 0)) / 1000, 5)
                return score

            best = sorted(candidate_rows, key=row_score, reverse=True)[0]
            rows.append(best)
        else:
            blank_streak += 1
            if blank_streak >= 5 and rows:
                break

    return rows

def deduplicate_rows_fuzzy(rows):
    """
    Fuzzy deduplication for mixed text/OCR extraction.
    """
    clean_rows = []
    seen = set()

    for row in rows:
        product = standardize_product_name(row.get("Product", ""))
        qty = int(row.get("Quantity", 0) or 0)
        amount = round(float(row.get("Amount", 0) or 0), 0)
        invoice = row.get("Invoice No", "")
        key = (invoice, product.lower(), qty, amount)

        if key not in seen:
            seen.add(key)
            row["Product"] = product
            clean_rows.append(row)

    return clean_rows


def parse_invoice_pdf_with_ocr(uploaded_file, metadata):
    """
    OCR fallback for scanned PDFs. Uses grid/table OCR first.
    Full-page OCR is avoided unless table-grid extraction finds nothing, because it is slower and less accurate.
    """
    if not OCR_AVAILABLE:
        return [], "OCR libraries are not installed. Install pytesseract, pillow, opencv-python-headless and tesseract-ocr."

    rows = []
    try:
        page_images = render_pdf_pages_for_ocr(uploaded_file, zoom=4)
        for image in page_images:
            grid_rows = parse_rows_from_ocr_grid(image, metadata)
            rows.extend(grid_rows)

            # Slow fallback only when grid detection completely fails.
            if not grid_rows:
                rows.extend(parse_rows_from_ocr_full_page(image, metadata))

        rows = deduplicate_rows_fuzzy(rows)
        return rows, ""
    except Exception as exc:
        return [], f"OCR failed: {exc}"

def auto_width(ws, min_width=10, max_width=42):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def style_data_sheet(ws, money_columns=None, number_columns=None, date_columns=None):
    money_columns = money_columns or []
    number_columns = number_columns or []
    date_columns = date_columns or []

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    headers = {}
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        headers[cell.value] = cell.column

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    for col_name in money_columns:
        if col_name in headers:
            col_letter = get_column_letter(headers[col_name])
            for cell in ws[col_letter][1:]:
                cell.number_format = '"Rs." #,##0'

    for col_name in number_columns:
        if col_name in headers:
            col_letter = get_column_letter(headers[col_name])
            for cell in ws[col_letter][1:]:
                cell.number_format = '#,##0'

    for col_name in date_columns:
        if col_name in headers:
            col_letter = get_column_letter(headers[col_name])
            for cell in ws[col_letter][1:]:
                cell.number_format = 'yyyy-mm-dd'

    ws.row_dimensions[1].height = 24
    auto_width(ws)


def create_dashboard_sheet(workbook, filtered_df, product_summary, invoice_summary, summary_text):
    if "Dashboard" in workbook.sheetnames:
        ws = workbook["Dashboard"]
    else:
        ws = workbook.create_sheet("Dashboard", 0)

    ws.sheet_view.showGridLines = False

    dark_fill = PatternFill("solid", fgColor="0F172A")
    blue_fill = PatternFill("solid", fgColor="1E3A8A")
    light_fill = PatternFill("solid", fgColor="EFF6FF")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=18)
    label_font = Font(color="1E3A8A", bold=True)
    value_font = Font(color="0F172A", bold=True, size=14)
    border_side = Side(style="thin", color="93C5FD")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    ws.merge_cells("A1:H2")
    ws["A1"] = "Invoice Sales Analytics Dashboard"
    ws["A1"].fill = dark_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for row in range(1, 3):
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = dark_fill

    total_sales = filtered_df["Amount"].sum()
    total_qty = filtered_df["Quantity"].sum()
    total_invoices = filtered_df["Invoice No"].nunique()
    total_products = filtered_df["Product"].nunique()
    total_discount = filtered_df.groupby("Invoice No")["Benefit / Discount"].max().sum()
    total_freight = filtered_df.groupby("Invoice No")["Freight Fee"].max().sum()
    total_balance = filtered_df.groupby("Invoice No")["Invoice Balance"].max().sum()
    avg_invoice_value = invoice_summary["Amount"].mean() if not invoice_summary.empty else 0

    kpis = [
        ("Total Sales", total_sales),
        ("Total Quantity", total_qty),
        ("Invoices", total_invoices),
        ("Unique Products", total_products),
        ("Average Invoice Value", avg_invoice_value),
        ("Net Balance", total_balance),
        ("Discount / Benefit", total_discount),
        ("Freight Fee", total_freight),
    ]

    start_row = 4
    for idx, (label, value) in enumerate(kpis):
        r = start_row + (idx // 4) * 3
        c = 1 + (idx % 4) * 2
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+1)
        ws.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c+1)
        ws.cell(r, c).value = label
        ws.cell(r, c).fill = blue_fill
        ws.cell(r, c).font = white_font
        ws.cell(r, c).alignment = Alignment(horizontal="center")
        ws.cell(r+1, c).value = value
        ws.cell(r+1, c).fill = light_fill
        ws.cell(r+1, c).font = value_font
        ws.cell(r+1, c).alignment = Alignment(horizontal="center")
        if "Sales" in label or "Value" in label or "Balance" in label or "Discount" in label or "Freight" in label:
            ws.cell(r+1, c).number_format = '"Rs." #,##0'
        else:
            ws.cell(r+1, c).number_format = '#,##0'

        for rr in [r, r+1]:
            for cc in [c, c+1]:
                ws.cell(rr, cc).border = border

    ws.merge_cells("A11:H14")
    ws["A11"] = summary_text
    ws["A11"].fill = PatternFill("solid", fgColor="F8FAFC")
    ws["A11"].font = Font(color="0F172A", size=11)
    ws["A11"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A16"] = "Top Products by Sales"
    ws["A16"].font = label_font

    top5 = product_summary.sort_values("Amount", ascending=False).head(5)
    headers = ["Product", "Quantity", "Amount", "Invoices"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(17, col_idx, header)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, (_, row) in enumerate(top5.iterrows(), 18):
        ws.cell(row_idx, 1, row["Product"])
        ws.cell(row_idx, 2, row["Quantity"])
        ws.cell(row_idx, 3, row["Amount"])
        ws.cell(row_idx, 4, row["Invoices"])
        for col_idx in range(1, 5):
            ws.cell(row_idx, col_idx).border = border
            ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.cell(row_idx, 2).number_format = '#,##0'
        ws.cell(row_idx, 3).number_format = '"Rs." #,##0'
        ws.cell(row_idx, 4).number_format = '#,##0'

    ws["F16"] = "Low-Selling Products by Quantity"
    ws["F16"].font = label_font

    bottom5 = product_summary[product_summary["Quantity"] > 0].sort_values("Quantity", ascending=True).head(5)
    for col_idx, header in enumerate(headers, 6):
        cell = ws.cell(17, col_idx, header)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_idx, (_, row) in enumerate(bottom5.iterrows(), 18):
        ws.cell(row_idx, 6, row["Product"])
        ws.cell(row_idx, 7, row["Quantity"])
        ws.cell(row_idx, 8, row["Amount"])
        ws.cell(row_idx, 9, row["Invoices"])
        for col_idx in range(6, 10):
            ws.cell(row_idx, col_idx).border = border
            ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.cell(row_idx, 7).number_format = '#,##0'
        ws.cell(row_idx, 8).number_format = '"Rs." #,##0'
        ws.cell(row_idx, 9).number_format = '#,##0'

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[11].height = 72


def convert_report_to_excel(df, product_summary, invoice_summary, summary_text):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        product_summary.to_excel(writer, index=False, sheet_name="Product Summary")
        invoice_summary.to_excel(writer, index=False, sheet_name="Invoice Summary")
        df.to_excel(writer, index=False, sheet_name="Extracted Rows")

        top5_amount = product_summary.sort_values("Amount", ascending=False).head(5)
        top5_qty = product_summary.sort_values("Quantity", ascending=False).head(5)
        bottom5_amount = product_summary[product_summary["Amount"] > 0].sort_values("Amount", ascending=True).head(5)
        bottom5_qty = product_summary[product_summary["Quantity"] > 0].sort_values("Quantity", ascending=True).head(5)

        top5_amount.to_excel(writer, index=False, sheet_name="Top 5 Sales")
        top5_qty.to_excel(writer, index=False, sheet_name="Top 5 Quantity")
        bottom5_amount.to_excel(writer, index=False, sheet_name="Bottom 5 Sales")
        bottom5_qty.to_excel(writer, index=False, sheet_name="Bottom 5 Quantity")

        workbook = writer.book

        create_dashboard_sheet(workbook, df, product_summary, invoice_summary, summary_text)

        money_cols = [
            "Amount", "Invoice_Subtotal", "Invoice_Balance", "Freight_Fee",
            "Benefit_Discount", "Invoice Subtotal", "Invoice Balance",
            "Freight Fee", "Benefit / Discount", "Calculated Amount",
            "Amount Difference", "Carton / Unit Price"
        ]
        number_cols = ["Quantity", "Invoices", "Sr No"]
        date_cols = ["Invoice Date", "Due Date"]

        for sheet_name in workbook.sheetnames:
            if sheet_name == "Dashboard":
                continue
            ws = workbook[sheet_name]
            style_data_sheet(ws, money_columns=money_cols, number_columns=number_cols, date_columns=date_cols)

    output.seek(0)
    return output


# -----------------------------
# Chart images for PDF
# -----------------------------
def make_matplotlib_bar(df, label_col, value_col, title, horizontal=True, max_rows=10):
    chart_df = df.head(max_rows).copy()
    fig, ax = plt.subplots(figsize=(8.8, 4.2), dpi=160)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#f8fafc")

    labels = chart_df[label_col].astype(str).tolist()
    values = chart_df[value_col].astype(float).tolist()

    if horizontal:
        labels = labels[::-1]
        values = values[::-1]
        ax.barh(labels, values, color="#2563eb")
        ax.set_xlabel(value_col)
    else:
        ax.bar(labels, values, color="#2563eb")
        ax.tick_params(axis="x", rotation=35)
        ax.set_ylabel(value_col)

    ax.set_title(title, fontsize=13, fontweight="bold", color="#0f172a")
    ax.grid(axis="x" if horizontal else "y", linestyle="--", alpha=0.35)
    ax.spines[["top", "right"]].set_visible(False)
    plt.tight_layout()

    img = BytesIO()
    fig.savefig(img, format="png", bbox_inches="tight")
    plt.close(fig)
    img.seek(0)
    return img


def make_matplotlib_pie(df, label_col, value_col, title, max_rows=8):
    chart_df = df.head(max_rows).copy()
    fig, ax = plt.subplots(figsize=(7.2, 4.4), dpi=160)
    fig.patch.set_facecolor("white")

    labels = chart_df[label_col].astype(str).tolist()
    values = chart_df[value_col].astype(float).tolist()

    wedges, texts, autotexts = ax.pie(
        values,
        labels=labels,
        autopct="%1.1f%%",
        startangle=90,
        pctdistance=0.78,
        textprops={"fontsize": 8},
    )
    centre_circle = plt.Circle((0, 0), 0.50, fc="white")
    fig.gca().add_artist(centre_circle)
    ax.set_title(title, fontsize=13, fontweight="bold", color="#0f172a")
    plt.tight_layout()

    img = BytesIO()
    fig.savefig(img, format="png", bbox_inches="tight")
    plt.close(fig)
    img.seek(0)
    return img


def make_matplotlib_line(df, x_col, y_col, title):
    chart_df = df.copy()
    fig, ax = plt.subplots(figsize=(8.8, 4.0), dpi=160)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#f8fafc")

    ax.plot(chart_df[x_col].astype(str), chart_df[y_col].astype(float), marker="o", linewidth=2.5, color="#2563eb")
    ax.set_title(title, fontsize=13, fontweight="bold", color="#0f172a")
    ax.set_ylabel(y_col)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.spines[["top", "right"]].set_visible(False)
    plt.xticks(rotation=25)
    plt.tight_layout()

    img = BytesIO()
    fig.savefig(img, format="png", bbox_inches="tight")
    plt.close(fig)
    img.seek(0)
    return img


# -----------------------------
# PDF report
# -----------------------------
def make_report_table(df, columns, widths=None):
    data = [columns]
    for _, row in df[columns].iterrows():
        data.append([str(row[col]) for col in columns])

    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def add_image_row(elements, img1, img2=None, width=4.7*inch, height=2.45*inch):
    if img2 is None:
        elements.append(Image(img1, width=7.8*inch, height=3.7*inch))
    else:
        table = Table(
            [[Image(img1, width=width, height=height), Image(img2, width=width, height=height)]],
            colWidths=[width + 0.15*inch, width + 0.15*inch]
        )
        table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ]))
        elements.append(table)


def create_pdf_dashboard_report(filtered_df, product_summary, invoice_summary, monthly_sales, summary_text):
    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=0.40 * inch,
        leftMargin=0.40 * inch,
        topMargin=0.40 * inch,
        bottomMargin=0.40 * inch,
        title="Invoice Sales Dashboard Report"
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        fontSize=20,
        textColor=colors.HexColor("#0f172a"),
        spaceAfter=14,
    )
    heading_style = ParagraphStyle(
        "Heading",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=13,
        textColor=colors.HexColor("#1e3a8a"),
        spaceBefore=10,
        spaceAfter=8,
    )
    normal_style = ParagraphStyle(
        "NormalCustom",
        parent=styles["Normal"],
        alignment=TA_LEFT,
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#0f172a"),
    )

    elements = []
    elements.append(Paragraph("Invoice Sales Dashboard Report", title_style))

    total_sales = filtered_df["Amount"].sum()
    total_qty = filtered_df["Quantity"].sum()
    total_invoices = filtered_df["Invoice No"].nunique()
    total_products = filtered_df["Product"].nunique()
    total_discount = filtered_df.groupby("Invoice No")["Benefit / Discount"].max().sum()
    total_freight = filtered_df.groupby("Invoice No")["Freight Fee"].max().sum()
    total_balance = filtered_df.groupby("Invoice No")["Invoice Balance"].max().sum()
    avg_invoice_value = invoice_summary["Amount"].mean() if not invoice_summary.empty else 0

    kpi_df = pd.DataFrame([
        ["Total Sales", format_currency(total_sales), "Total Quantity", format_number(total_qty)],
        ["Invoices", format_number(total_invoices), "Unique Products", format_number(total_products)],
        ["Average Invoice Value", format_currency(avg_invoice_value), "Net Balance", format_currency(total_balance)],
        ["Total Discount / Benefit", format_currency(total_discount), "Total Freight Fee", format_currency(total_freight)],
    ])

    elements.append(Paragraph("Key Performance Indicators", heading_style))
    kpi_table = Table(kpi_df.values.tolist(), colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#93c5fd")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 10))

    elements.append(Paragraph("Automatic Business Summary", heading_style))
    elements.append(Paragraph(summary_text, normal_style))
    elements.append(Spacer(1, 8))

    # Static chart images included in PDF
    top_sales_img = make_matplotlib_bar(product_summary.sort_values("Amount", ascending=False), "Product", "Amount", "Top 10 Products by Sales")
    sales_share_img = make_matplotlib_pie(product_summary.sort_values("Amount", ascending=False), "Product", "Amount", "Sales Share by Product")
    top_qty_img = make_matplotlib_bar(product_summary.sort_values("Quantity", ascending=False), "Product", "Quantity", "Top 10 Products by Quantity")
    bottom_qty_img = make_matplotlib_bar(product_summary[product_summary["Quantity"] > 0].sort_values("Quantity", ascending=True), "Product", "Quantity", "Bottom 5 Products by Quantity", max_rows=5)
    sales_trend_img = make_matplotlib_line(monthly_sales, "Month", "Amount", "Monthly Sales Trend")
    qty_trend_img = make_matplotlib_line(monthly_sales, "Month", "Quantity", "Monthly Quantity Trend")

    elements.append(Paragraph("Dashboard Charts", heading_style))
    add_image_row(elements, top_sales_img, sales_share_img)
    elements.append(Spacer(1, 6))
    add_image_row(elements, top_qty_img, bottom_qty_img)
    elements.append(PageBreak())

    elements.append(Paragraph("Monthly Trends", heading_style))
    add_image_row(elements, sales_trend_img, qty_trend_img)

    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Top and Low-Selling Product Tables", heading_style))

    top5_amount = product_summary.sort_values("Amount", ascending=False).head(5).copy()
    bottom5_qty = product_summary[product_summary["Quantity"] > 0].sort_values("Quantity", ascending=True).head(5).copy()

    for table_df in [top5_amount, bottom5_qty]:
        table_df["Amount"] = table_df["Amount"].apply(format_currency)
        table_df["Quantity"] = table_df["Quantity"].apply(format_number)

    product_tables = Table(
        [[
            make_report_table(top5_amount, ["Product", "Quantity", "Amount", "Invoices"],
                              widths=[2.2*inch, 0.9*inch, 1.1*inch, 0.7*inch]),
            make_report_table(bottom5_qty, ["Product", "Quantity", "Amount", "Invoices"],
                              widths=[2.2*inch, 0.9*inch, 1.1*inch, 0.7*inch])
        ]],
        colWidths=[5.1*inch, 5.1*inch]
    )
    product_tables.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(product_tables)

    elements.append(PageBreak())

    invoice_pdf = invoice_summary.head(30).copy()
    for col in ["Amount", "Invoice_Subtotal", "Invoice_Balance", "Freight_Fee", "Benefit_Discount"]:
        if col in invoice_pdf.columns:
            invoice_pdf[col] = invoice_pdf[col].apply(format_currency)
    invoice_pdf["Quantity"] = invoice_pdf["Quantity"].apply(format_number)
    invoice_pdf["Invoice Date"] = invoice_pdf["Invoice Date"].astype(str)

    elements.append(Paragraph("Invoice Summary", heading_style))
    elements.append(make_report_table(
        invoice_pdf,
        ["Invoice No", "Invoice Date", "Quantity", "Amount", "Invoice_Balance", "Freight_Fee", "Benefit_Discount"],
        widths=[1.2*inch, 1.4*inch, 1.0*inch, 1.3*inch, 1.3*inch, 1.1*inch, 1.3*inch]
    ))

    doc.build(elements)
    output.seek(0)
    return output


# -----------------------------
# Streamlit UI
# -----------------------------
uploaded_files = st.file_uploader(
    "Upload invoice PDF or image files",
    type=["pdf", "png", "jpg", "jpeg", "webp", "bmp", "tif", "tiff"],
    accept_multiple_files=True
)

with st.expander("App features in this version"):
    st.write(
        """
        ✅ Higher-accuracy OCR fallback added for scanned / image-based PDFs  
        ✅ OCR review/correction table added before dashboard calculations  
        ✅ Premium CSS styling applied to dashboard, sidebar, KPI cards, uploader, buttons, tabs and charts  
        ✅ Invoice filter retained  
        ✅ AI-style automatic business summary retained  
        ✅ PDF report now includes charts  
        ✅ Excel download now has professional styling, dashboard sheet, filters, frozen headers and number formats  
        """
    )

if uploaded_files:
    all_rows = []
    debug_text = {}
    debug_lines = {}
    ocr_messages = {}

    for uploaded_file in uploaded_files:
        rows, text, visual_lines, ocr_message = parse_uploaded_invoice_file(uploaded_file)
        all_rows.extend(rows)
        debug_text[uploaded_file.name] = text
        debug_lines[uploaded_file.name] = visual_lines
        if ocr_message:
            ocr_messages[uploaded_file.name] = ocr_message

    if not all_rows:
        st.error("No invoice table rows found.")
        st.warning("Open the Debug section to check how the PDF text is being read. If this is a scanned/image-based PDF, OCR dependencies may be missing.")
        with st.expander("Debug"):
            if ocr_messages:
                st.write("OCR messages:", ocr_messages)
            for file_name, text in debug_text.items():
                st.subheader(file_name)
                st.text_area("Extracted text", text, height=300)
                st.text_area("Visual lines", "\n".join(debug_lines[file_name]), height=300)

    else:
        df = pd.DataFrame(all_rows)
        df["Invoice Date"] = pd.to_datetime(df["Invoice Date"], errors="coerce")
        df["Due Date"] = pd.to_datetime(df["Due Date"], errors="coerce")
        df["Month"] = df["Invoice Date"].dt.to_period("M").astype(str)

        # OCR review/correction step: scanned PDFs can misread numbers.
        # The dashboard will use the corrected values from this editor.
        ocr_used = "Parser Method" in df.columns and df["Parser Method"].astype(str).str.contains("ocr", case=False, na=False).any()
        if ocr_used:
            st.warning(
                "Scanned/image-based PDF detected. OCR has been used. Please review the extracted values below; "
                "you can correct Product, Unit Price, Quantity, and Amount before the dashboard is calculated."
            )
            review_columns = [
                "Invoice No", "Invoice Date", "Product", "Carton / Unit Price", "Quantity", "Amount",
                "OCR Review Required", "OCR Auto Correction", "OCR Raw Text"
            ]
            available_review_columns = [col for col in review_columns if col in df.columns]
            with st.expander("Review / Correct OCR Extracted Rows", expanded=True):
                edited_review = st.data_editor(
                    df[available_review_columns],
                    use_container_width=True,
                    hide_index=True,
                    num_rows="dynamic",
                    key="ocr_review_editor"
                )
                editable_cols = ["Product", "Carton / Unit Price", "Quantity", "Amount"]
                for col in editable_cols:
                    if col in edited_review.columns:
                        df.loc[edited_review.index, col] = edited_review[col]

                for col in ["Carton / Unit Price", "Quantity", "Amount"]:
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

                if "Quantity" in df.columns:
                    df["Quantity"] = df["Quantity"].astype(int)
                if "Carton / Unit Price" in df.columns and "Quantity" in df.columns:
                    df["Calculated Amount"] = df["Carton / Unit Price"] * df["Quantity"]
                if "Calculated Amount" in df.columns and "Amount" in df.columns:
                    df["Amount Difference"] = df["Amount"] - df["Calculated Amount"]


        # Sidebar filters
        st.sidebar.header("Filters")

        invoices = sorted(df["Invoice No"].dropna().unique())
        selected_invoices = st.sidebar.multiselect("Invoice No", invoices, default=invoices)

        months = sorted([m for m in df["Month"].dropna().unique() if m != "NaT"])
        selected_months = st.sidebar.multiselect("Month", months, default=months)

        products = sorted(df["Product"].dropna().unique())
        selected_products = st.sidebar.multiselect("Products", products, default=products)

        filtered_df = df.copy()

        if selected_invoices:
            filtered_df = filtered_df[filtered_df["Invoice No"].isin(selected_invoices)]
        if selected_months:
            filtered_df = filtered_df[filtered_df["Month"].isin(selected_months)]
        if selected_products:
            filtered_df = filtered_df[filtered_df["Product"].isin(selected_products)]

        if filtered_df.empty:
            st.warning("No data available for selected filters.")
            st.stop()

        product_summary = build_product_summary(filtered_df)
        invoice_summary = build_invoice_summary(filtered_df)
        monthly_sales = (
            filtered_df.groupby("Month", as_index=False)
            .agg(Amount=("Amount", "sum"), Quantity=("Quantity", "sum"))
            .sort_values("Month")
        )

        total_sales = filtered_df["Amount"].sum()
        total_qty = filtered_df["Quantity"].sum()
        total_invoices = filtered_df["Invoice No"].nunique()
        total_products = filtered_df["Product"].nunique()
        total_discount = filtered_df.groupby("Invoice No")["Benefit / Discount"].max().sum()
        total_freight = filtered_df.groupby("Invoice No")["Freight Fee"].max().sum()
        total_balance = filtered_df.groupby("Invoice No")["Invoice Balance"].max().sum()
        avg_invoice_value = invoice_summary["Amount"].mean() if not invoice_summary.empty else 0

        top_revenue_product = product_summary.sort_values("Amount", ascending=False).iloc[0]["Product"]
        top_qty_product = product_summary.sort_values("Quantity", ascending=False).iloc[0]["Product"]

        bottom_qty_product = (
            product_summary[product_summary["Quantity"] > 0]
            .sort_values("Quantity", ascending=True)
            .iloc[0]["Product"]
        )
        bottom_revenue_product = (
            product_summary[product_summary["Amount"] > 0]
            .sort_values("Amount", ascending=True)
            .iloc[0]["Product"]
        )

        summary_text = generate_auto_summary(filtered_df, product_summary, invoice_summary, monthly_sales)

        st.success(f"Extracted {len(df)} sales row(s) from {len(uploaded_files)} PDF file(s).")

        if "Parser Method" in df.columns and df["Parser Method"].astype(str).str.contains("ocr", case=False, na=False).any():
            st.info(
                "OCR mode was used for at least one scanned/image-based PDF. "
                "Please review Product Summary / Excel output because OCR can misread numbers or product names."
            )


        # KPI row 1
        r1c1, r1c2, r1c3, r1c4 = st.columns(4)
        with r1c1:
            kpi_card("Total Sales", format_currency(total_sales), "Product-wise sales amount")
        with r1c2:
            kpi_card("Total Quantity", format_number(total_qty), "Total cartons / units sold")
        with r1c3:
            kpi_card("Invoices", format_number(total_invoices), "Filtered invoice count")
        with r1c4:
            kpi_card("Unique Products", format_number(total_products), "Different products sold")

        # KPI row 2
        r2c1, r2c2, r2c3, r2c4 = st.columns(4)
        with r2c1:
            kpi_card("Top Revenue Product", top_revenue_product, "Highest sales amount")
        with r2c2:
            kpi_card("Top Quantity Product", top_qty_product, "Highest quantity sold")
        with r2c3:
            kpi_card("Average Invoice Value", format_currency(avg_invoice_value), "Average sales per invoice")
        with r2c4:
            kpi_card("Net Balance", format_currency(total_balance), "Sum of filtered invoice balances")

        # KPI row 3
        r3c1, r3c2, r3c3, r3c4 = st.columns(4)
        with r3c1:
            kpi_card("Total Discount / Benefit", format_currency(total_discount), "Across unique invoices")
        with r3c2:
            kpi_card("Total Freight Fee", format_currency(total_freight), "Across unique invoices")
        with r3c3:
            kpi_card("Lowest Quantity Item", bottom_qty_product, "Low movement item")
        with r3c4:
            kpi_card("Lowest Revenue Item", bottom_revenue_product, "Low value item")

        st.subheader("🤖 AI Summary")
        st.markdown(f"<div class='summary-box'>{summary_text}</div>", unsafe_allow_html=True)

        st.divider()

        # Charts only. Removed the four table blocks from dashboard.
        ch1, ch2 = st.columns(2)
        with ch1:
            st.plotly_chart(
                styled_bar_chart(product_summary.head(10), "Product", "Amount", "Top 10 Products by Sales"),
                use_container_width=True
            )
        with ch2:
            st.plotly_chart(
                styled_pie_chart(product_summary.head(8), "Product", "Amount", "Sales Share by Product"),
                use_container_width=True
            )

        ch3, ch4 = st.columns(2)
        with ch3:
            qty_summary = (
                filtered_df.groupby("Product", as_index=False)["Quantity"]
                .sum()
                .sort_values("Quantity", ascending=False)
            )
            st.plotly_chart(
                styled_bar_chart(qty_summary.head(10), "Product", "Quantity", "Top 10 Products by Quantity"),
                use_container_width=True
            )
        with ch4:
            bottom5_qty = product_summary[product_summary["Quantity"] > 0].sort_values("Quantity", ascending=True).head(5)
            st.plotly_chart(
                styled_bar_chart(bottom5_qty, "Product", "Quantity", "Bottom 5 Products by Quantity"),
                use_container_width=True
            )

        ch5, ch6 = st.columns(2)
        with ch5:
            st.plotly_chart(
                styled_line_chart(monthly_sales, "Month", "Amount", "Monthly Sales Trend"),
                use_container_width=True
            )
        with ch6:
            st.plotly_chart(
                styled_line_chart(monthly_sales, "Month", "Quantity", "Monthly Quantity Trend"),
                use_container_width=True
            )

        # Tabs for data and downloads
        tab1, tab2, tab3, tab4 = st.tabs(["Product Summary", "Invoice Summary", "Export Reports", "Debug"])

        with tab1:
            st.subheader("Product Summary")
            display_product_summary = dataframe_for_display(
                product_summary,
                money_cols=["Amount"],
                number_cols=["Quantity", "Invoices"]
            )
            st.dataframe(display_product_summary, use_container_width=True, hide_index=True)

        with tab2:
            st.subheader("Invoice Summary")
            display_invoice_summary = invoice_summary.copy()
            display_invoice_summary["Invoice Date"] = display_invoice_summary["Invoice Date"].astype(str)
            display_invoice_summary = dataframe_for_display(
                display_invoice_summary,
                money_cols=["Amount", "Invoice_Subtotal", "Invoice_Balance", "Freight_Fee", "Benefit_Discount"],
                number_cols=["Quantity"]
            )
            st.dataframe(display_invoice_summary, use_container_width=True, hide_index=True)

        with tab3:
            st.subheader("Export Reports")

            col_pdf, col_excel = st.columns(2)

            with col_pdf:
                pdf_report = create_pdf_dashboard_report(
                    filtered_df,
                    product_summary,
                    invoice_summary,
                    monthly_sales,
                    summary_text
                )
                st.download_button(
                    label="📄 Download Dashboard PDF Report with Charts",
                    data=pdf_report,
                    file_name="invoice_dashboard_report_with_charts.pdf",
                    mime="application/pdf"
                )
                st.caption("PDF includes KPIs, AI summary, charts, top/bottom product tables and invoice summary.")

            with col_excel:
                excel_file = convert_report_to_excel(
                    filtered_df,
                    product_summary,
                    invoice_summary,
                    summary_text
                )
                st.download_button(
                    label="📥 Download Styled Excel Report",
                    data=excel_file,
                    file_name="styled_invoice_sales_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.caption("Excel includes a styled Dashboard sheet, filters, frozen headers and formatted summaries.")

        with tab4:
            st.subheader("Debug Tools")
            show_debug = st.checkbox("Show extracted text and reconstructed PDF lines")
            if show_debug:
                for file_name in debug_text:
                    st.markdown(f"**{file_name}**")
                    st.text_area("Extracted full text", debug_text[file_name], height=250)
                    st.text_area("Reconstructed visual lines", "\n".join(debug_lines[file_name]), height=250)

else:
    st.info("Upload one or more invoice PDF or image files to start.")